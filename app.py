import os
import sys
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scraper import get_paa_questions


def _base_path():
    """PyInstaller frozen exe ise _MEIPASS, değilse script dizini."""
    if getattr(sys, "frozen", False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _base_path()

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
)

# outputs klasörünü exe'nin yanına koy (MEIPASS değil, gerçek konum)
if getattr(sys, "frozen", False):
    _exe_dir = os.path.dirname(sys.executable)
else:
    _exe_dir = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(_exe_dir, "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/search", methods=["POST"])
def search():
    data = request.get_json()
    query = (data or {}).get("query", "").strip()
    num_questions = int((data or {}).get("num_questions", 10))
    num_questions = max(1, min(num_questions, 20))

    if not query:
        return jsonify({"error": "Sorgu boş olamaz."}), 400

    try:
        questions = get_paa_questions(query, num_questions=num_questions)
    except Exception as e:
        return jsonify({"error": f"Scraper hatası: {str(e)}"}), 500

    if not questions:
        return jsonify({
            "error": (
                "Bu sorgu için Google'da 'People Also Ask' soruları bulunamadı. "
                "Farklı bir sorgu deneyin."
            )
        }), 404

    try:
        filename = _create_excel(query, questions)
    except Exception as e:
        return jsonify({"error": f"Excel oluşturma hatası: {str(e)}"}), 500

    return jsonify({"questions": questions, "filename": filename})


@app.route("/bulk-search", methods=["POST"])
def bulk_search():
    data = request.get_json()
    queries = (data or {}).get("queries", [])
    num_questions = int((data or {}).get("num_questions", 5))
    num_questions = max(1, min(num_questions, 20))

    if not queries:
        return jsonify({"error": "Sorgu listesi boş."}), 400

    results = {}
    for query in queries:
        query = query.strip()
        if not query:
            continue
        try:
            questions = get_paa_questions(query, num_questions=num_questions)
            results[query] = questions
        except Exception:
            results[query] = []

    if not results:
        return jsonify({"error": "Hiçbir sorgu işlenemedi."}), 500

    try:
        filename = _create_bulk_excel(results)
    except Exception as e:
        return jsonify({"error": f"Excel hatası: {str(e)}"}), 500

    return jsonify({"results": results, "filename": filename})


@app.route("/bulk-search-excel", methods=["POST"])
def bulk_search_excel():
    data = request.get_json()
    results = (data or {}).get("results", {})

    if not results:
        return jsonify({"error": "Sonuç verisi boş."}), 400

    try:
        filename = _create_bulk_excel(results)
    except Exception as e:
        return jsonify({"error": f"Excel hatası: {str(e)}"}), 500

    return jsonify({"filename": filename})


@app.route("/update-excel", methods=["POST"])
def update_excel():
    data = request.get_json()
    query = (data or {}).get("query", "").strip()
    questions = (data or {}).get("questions", [])

    if not query or not questions:
        return jsonify({"error": "Veri eksik."}), 400

    try:
        filename = _create_excel(query, questions)
    except Exception as e:
        return jsonify({"error": f"Excel hatası: {str(e)}"}), 500

    return jsonify({"filename": filename})


@app.route("/download/<path:filename>")
def download(filename):
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(filepath):
        return jsonify({"error": "Dosya bulunamadı."}), 404
    return send_file(filepath, as_attachment=True)


def _create_excel(query, questions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAA Sonuçları"

    # --- Stiller ---
    blue = "1A73E8"
    light_blue = "E8F0FE"
    white = "FFFFFF"
    dark = "202124"

    header_font = Font(bold=True, color=white, size=11)
    title_font = Font(bold=True, size=13, color=dark)
    data_font = Font(size=11, color=dark)

    blue_fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
    alt_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Satır 1: Başlık ---
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Google People Also Ask — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    # --- Satır 2: Kolon başlıkları ---
    ws["A2"] = "Sorgu"
    ws["B2"] = "People Also Ask Soruları"
    for cell in [ws["A2"], ws["B2"]]:
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 28

    # --- Veri satırları ---
    for i, question in enumerate(questions):
        row = i + 3
        ws[f"A{row}"] = query if i == 0 else ""
        ws[f"B{row}"] = question

        fill = alt_fill if i % 2 == 1 else None
        for col in ["A", "B"]:
            cell = ws[f"{col}{row}"]
            cell.font = data_font
            cell.alignment = left_wrap
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row].height = 38

    # Sorgu sütununu birleştir
    if len(questions) > 1:
        ws.merge_cells(f"A3:A{len(questions) + 2}")
        ws["A3"].alignment = center

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 65

    # Dosya adı
    safe_query = "".join(
        c for c in query if c.isalnum() or c in (" ", "-", "_")
    ).strip()[:30]
    filename = f"PAA_{safe_query}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    return filename


def _create_bulk_excel(results):
    """Toplu sorgu sonuçlarını tek bir Excel dosyasına yaz."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAA Toplu Sonuçlar"

    blue = "1A73E8"
    light_blue = "E8F0FE"
    white = "FFFFFF"
    dark = "202124"

    header_font = Font(bold=True, color=white, size=11)
    title_font = Font(bold=True, size=13, color=dark)
    data_font = Font(size=11, color=dark)
    query_font = Font(bold=True, size=11, color=dark)

    blue_fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
    alt_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Başlık
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Google PAA Toplu Sorgu — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    # Kolon başlıkları
    ws["A2"] = "Sorgu"
    ws["B2"] = "People Also Ask Soruları"
    for cell in [ws["A2"], ws["B2"]]:
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 28

    row = 3
    for query, questions in results.items():
        if not questions:
            ws[f"A{row}"] = query
            ws[f"B{row}"] = "(Sonuç bulunamadı)"
            for col in ["A", "B"]:
                cell = ws[f"{col}{row}"]
                cell.font = data_font
                cell.alignment = left_wrap
                cell.border = border
            ws.row_dimensions[row].height = 38
            row += 1
            continue

        start_row = row
        for i, question in enumerate(questions):
            ws[f"A{row}"] = query if i == 0 else ""
            ws[f"B{row}"] = question

            fill = alt_fill if i % 2 == 1 else None
            for col in ["A", "B"]:
                cell = ws[f"{col}{row}"]
                cell.font = data_font
                cell.alignment = left_wrap
                cell.border = border
                if fill:
                    cell.fill = fill
            ws.row_dimensions[row].height = 38
            row += 1

        # Sorgu hücresini birleştir
        if len(questions) > 1:
            ws.merge_cells(f"A{start_row}:A{row - 1}")
        ws[f"A{start_row}"].font = query_font
        ws[f"A{start_row}"].alignment = center

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 65

    filename = f"PAA_Toplu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filename


if __name__ == "__main__":
    print("Sunucu başlatılıyor: http://localhost:8080")
    app.run(debug=False, port=8080)
